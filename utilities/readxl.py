import openpyxl

class ReadWriteExcel():

    def __init__(self):
        self.filename = "..\\data\\Data.xlsx"

    def getMaxRow(self,sheetName):
        workbook = openpyxl.load_workbook(self.filename)
        sheet = workbook[sheetName]
        return sheet.max_row

    def getMaxColmn(self,sheetName):
        workbook = openpyxl.load_workbook(self.filename)
        sheet = workbook[sheetName]
        return sheet.max_column

    def getCellValue(self,sheetName,rowNum, colNum):
        workbook = openpyxl.load_workbook(self.filename)
        sheet = workbook[sheetName]
        return sheet.cell(row=rowNum, column=colNum).value

    def setCellValue(self,sheetName,rowNum, colNum,data):
        workbook = openpyxl.load_workbook(self.filename)
        sheet = workbook[sheetName]
        sheet.cell(row=rowNum, column=colNum).value = data
        workbook.save(self.filename)


def getAlldata(sheetName):
    obj = ReadWriteExcel()
    rowValues = []
    rows = obj.getMaxRow(sheetName)
    colms = obj.getMaxColmn(sheetName)

    for i in range(2,rows+1):
        columValues = []
        for j in range(1,colms+1):
            cellValue = obj.getCellValue(sheetName,i,j)
            columValues.insert(j,cellValue)
        rowValues.insert(i,columValues)
    return rowValues
